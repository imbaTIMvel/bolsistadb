"""
DADOS BANCÁRIOS — Automação de tratamento de dados bancários dos alunos bolsistas
Baseado na proposta de automação e seguindo a arquitetura do SmartPC.
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from copy import copy
import unicodedata
import re
import os
import sys
import subprocess
from datetime import datetime, date

from PySide6.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel, QFileDialog,
    QMessageBox, QVBoxLayout, QHBoxLayout, QFrame, QScrollArea,
    QCheckBox
)
from PySide6.QtGui import QIcon, QPixmap, QPainter, QColor
from PySide6.QtCore import Qt, QThread, Signal, QRectF


# ==============================
# BANCO DE DADOS FEBRABAN
# Código → Nome canônico
# Fonte: lista FEBRABAN (principais bancos BR)
# ==============================

BANCOS_FEBRABAN = {
    "001": "Banco do Brasil S.A.",
    "003": "Banco da Amazônia S.A.",
    "004": "Banco do Nordeste do Brasil S.A.",
    "021": "Banestes S.A.",
    "025": "Banco Alfa S.A.",
    "033": "Banco Santander (Brasil) S.A.",
    "036": "Banco Bradesco BBI S.A.",
    "037": "Banco do Estado do Pará S.A.",
    "041": "Banco do Estado do Rio Grande do Sul S.A.",
    "047": "Banco do Estado de Sergipe S.A.",
    "062": "Hipercard Banco Múltiplo S.A.",
    "069": "Banco Crefisa S.A.",
    "070": "Banco de Brasília S.A. — BRB",
    "077": "Banco Inter S.A.",
    "084": "Uniprime Norte do Paraná",
    "085": "Cooperativa Central de Crédito Urbano — Cecred",
    "089": "Cooperativa de Crédito Rural da Região da Mogiana",
    "097": "Cooperativa Central de Crédito Noroeste Brasileiro",
    "099": "Uniprime Central",
    "104": "Caixa Econômica Federal",
    "107": "Banco Bocom BBM S.A.",
    "114": "Central Cooperativa de Crédito no Estado do Espírito Santo",
    "133": "Cresol Confederação",
    "136": "Unicred do Brasil",
    "197": "Stone Pagamentos S.A.",
    "208": "Banco BTG Pactual S.A.",
    "212": "Banco Original S.A.",
    "218": "Banco BS2 S.A.",
    "224": "Banco Fibra S.A.",
    "237": "Banco Bradesco S.A.",
    "243": "Banco Máxima S.A.",
    "246": "Banco ABC Brasil S.A.",
    "260": "Nu Pagamentos S.A. — Nubank",
    "290": "Pagseguro Internet S.A.",
    "318": "Banco BMG S.A.",
    "320": "China Construction Bank (Brasil) Banco Múltiplo S.A.",
    "336": "Banco C6 S.A.",
    "341": "Itaú Unibanco S.A.",
    "348": "Banco XP S.A.",
    "364": "Gerencianet Pagamentos do Brasil Ltda.",
    "376": "Banco J. P. Morgan S.A.",
    "389": "Banco Mercantil do Brasil S.A.",
    "394": "Banco Bradesco Financiamentos S.A.",
    "399": "Kirton Bank S.A.",
    "422": "Banco Safra S.A.",
    "456": "Banco MUFG Brasil S.A.",
    "473": "Banco Caixa Geral — Brasil S.A.",
    "477": "Citibank N.A.",
    "487": "Deutsche Bank S.A.",
    "505": "Banco Credit Suisse (Brasil) S.A.",
    "600": "Banco Luso Brasileiro S.A.",
    "604": "Banco Industrial do Brasil S.A.",
    "610": "Banco VR S.A.",
    "611": "Banco Paulista S.A.",
    "613": "Omni Banco S.A.",
    "623": "Banco Pan S.A.",
    "626": "Banco C6 Consignado S.A.",
    "630": "Banco Smartbank S.A.",
    "633": "Banco Rendimento S.A.",
    "634": "Banco Triângulo S.A.",
    "637": "Banco Sofisa S.A.",
    "643": "Banco Pine S.A.",
    "652": "Itaú Unibanco Holding S.A.",
    "653": "Banco Indusval S.A.",
    "654": "Banco A.J.Renner S.A.",
    "655": "Banco Votorantim S.A.",
    "707": "Banco Daycoval S.A.",
    "719": "Banif — Banco Internacional do Funchal (Brasil) S.A.",
    "735": "Banco Neon S.A.",
    "739": "Banco Cetelem S.A.",
    "741": "Banco Ribeirão Preto S.A.",
    "743": "Banco Semear S.A.",
    "745": "Banco Citibank S.A.",
    "746": "Banco Modal S.A.",
    "747": "Banco Rabobank International Brasil S.A.",
    "748": "Sicredi",
    "751": "Scotiabank Brasil S.A.",
    "752": "Banco BNP Paribas Brasil S.A.",
    "753": "Novo Banco Continental S.A.",
    "754": "Banco Sistema S.A.",
    "755": "Bank of America Merrill Lynch Banco Múltiplo S.A.",
    "756": "Bancoob / Sicoob",
    "757": "Banco KEB Hana do Brasil S.A.",
    "323": "Mercado Pago Instituição de Pagamento Ltda.",
    "380": "PicPay Serviços S.A.",
}

# Alias de nomes comuns para facilitar o fuzzy match
ALIASES_BANCOS = {
    "bb": "001",
    "banco do brasil": "001",
    "banese": "047",
    "banpara": "037",
    "banrisul": "041",
    "bradesco": "237",
    "btg": "208",
    "btg pactual": "208",
    "c6": "336",
    "c6 bank": "336",
    "caixa": "104",
    "cef": "104",
    "caixa economica": "104",
    "caixa economica federal": "104",
    "citibank": "477",
    "cresol": "133",
    "daycoval": "707",
    "inter": "077",
    "banco inter": "077",
    "itau": "341",
    "itaú": "341",
    "itau unibanco": "341",
    "nubank": "260",
    "nu": "260",
    "neon": "735",
    "original": "212",
    "pagbank": "290",
    "pagseguro": "290",
    "pan": "623",
    "santander": "033",
    "sicoob": "756",
    "bancoob": "756",
    "sicredi": "748",
    "stone": "197",
    "uniprime": "084",
    "votorantim": "655",
    "xp": "348",
    "xp investimentos": "348",
    "mercado pago": "323",
    "mercadopago": "323",
    "picpay": "380",
    "pic pay": "380",
}


# ==============================
# UTILITÁRIOS DE TEXTO
# ==============================

def normalizar_texto(texto):
    if texto is None or (isinstance(texto, float) and pd.isna(texto)):
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    texto = texto.replace("º", "o").replace("°", "o")
    return texto


def normalizar_nome(nome):
    """Normaliza nome para comparação: strip, upper, remove espaços duplos."""
    if nome is None or (isinstance(nome, float) and pd.isna(nome)):
        return ""
    nome = str(nome).strip().upper()
    nome = re.sub(r'\s+', ' ', nome)
    return nome


def normalizar_colunas(df):
    df.columns = [normalizar_texto(c) for c in df.columns]
    return df


def encontrar_coluna_por_keywords(df, keywords, excluir_keywords=None, obrigatoria=True):
    """
    Busca colunas que contenham TODAS as keywords fornecidas.
    Retorna a primeira coluna (não vazia) que satisfaz a condição.
    """
    colunas_norm = {normalizar_texto(c): c for c in df.columns}
    candidatos = []

    for col_norm, col_orig in colunas_norm.items():
        if all(kw in col_norm for kw in keywords):
            if excluir_keywords:
                if any(ekw in col_norm for ekw in excluir_keywords):
                    continue
            candidatos.append(col_orig)

    # Prioriza a primeira com pelo menos um valor não-vazio
    for col in candidatos:
        if df[col].notna().any() and (df[col].astype(str).str.strip() != "").any():
            return col

    # Se todas estão vazias, retorna a primeira mesmo assim
    if candidatos:
        return candidatos[0]

    if obrigatoria:
        raise Exception(f"Coluna com keywords {keywords} não encontrada.")
    return None


def ler_excel(path, **kwargs):
    ext = os.path.splitext(path)[1].lower()
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    return pd.read_excel(path, engine=engine, dtype=str, **kwargs)


# ==============================
# TRATAMENTO DE BANCO
# ==============================

def _similaridade_levenshtein(a, b):
    """Distância de Levenshtein simples para strings curtas."""
    if len(a) == 0: return len(b)
    if len(b) == 0: return len(a)
    matriz = [[0] * (len(b) + 1) for _ in range(len(a) + 1)]
    for i in range(len(a) + 1): matriz[i][0] = i
    for j in range(len(b) + 1): matriz[0][j] = j
    for i in range(1, len(a) + 1):
        for j in range(1, len(b) + 1):
            custo = 0 if a[i - 1] == b[j - 1] else 1
            matriz[i][j] = min(
                matriz[i - 1][j] + 1,
                matriz[i][j - 1] + 1,
                matriz[i - 1][j - 1] + custo
            )
    return matriz[len(a)][len(b)]


def _score_similaridade(a, b):
    """Retorna score 0..1 (1 = idêntico) usando Levenshtein."""
    dist = _similaridade_levenshtein(a, b)
    max_len = max(len(a), len(b), 1)
    return 1.0 - dist / max_len


def resolver_banco(valor_bruto):
    """
    Recebe entrada bruta do campo banco/código.
    Retorna (codigo, nome_canonico, confianca, alerta).

    Estratégias em cascata — priorizando identificação pelo NOME:
    1. Alias exato (nomes comuns e abreviações conhecidas)
    2. Fuzzy match contra nomes canônicos FEBRABAN (Levenshtein)
    3. Código numérico extraído da string (último recurso)
    4. Falha com confiança = 0

    A priorização por nome evita que um código numérico presente na string
    (mas referente a outra versão do formulário ou digitação errônea) prevaleça
    sobre o nome claramente legível do banco.
    """
    if not valor_bruto or normalizar_texto(valor_bruto) == "":
        return ("", "", 0.0, "Campo banco em branco")

    entrada = normalizar_texto(valor_bruto)

    # Palavras genéricas a ignorar no fuzzy
    _GENERICAS = r'\b(banco|s\.?a\.?|ltda|do|da|de|e|brasil|instituicao|pagamento|cooperativo|cooperativa)\b'

    # 1. Alias exato (inclui abreviações e nomes populares)
    for alias, codigo in ALIASES_BANCOS.items():
        if alias in entrada:
            return (codigo, BANCOS_FEBRABAN[codigo], 1.0, "")

    # 2. Fuzzy match contra nomes canônicos FEBRABAN
    entrada_limpa = re.sub(_GENERICAS, '', entrada).strip() or entrada
    melhor_score  = 0.0
    melhor_codigo = ""
    melhor_nome   = ""

    for codigo, nome in BANCOS_FEBRABAN.items():
        nome_limpo = re.sub(_GENERICAS, '', normalizar_texto(nome)).strip()
        score = _score_similaridade(entrada_limpa, nome_limpo)
        if score > melhor_score:
            melhor_score  = score
            melhor_codigo = codigo
            melhor_nome   = nome

    THRESHOLD_FUZZY = 0.60
    if melhor_score >= THRESHOLD_FUZZY:
        alerta = (f"Banco identificado por similaridade ({melhor_score:.0%}): verificar"
                  if melhor_score < 0.95 else "")
        return (melhor_codigo, melhor_nome, melhor_score, alerta)

    # 3. Código numérico (último recurso — apenas quando nome não identificou)
    codigos_encontrados = re.findall(r'\b(\d{3,4})\b', entrada)
    for codigo_raw in codigos_encontrados:
        for codigo in [codigo_raw.zfill(3), codigo_raw.lstrip('0').zfill(3)]:
            if codigo in BANCOS_FEBRABAN:
                return (codigo, BANCOS_FEBRABAN[codigo], 0.80,
                        "Identificado apenas pelo código numérico — confirmar nome")

    return ("", str(valor_bruto).strip(), 0.0,
            f"Banco não identificado na tabela FEBRABAN: '{valor_bruto}'")


# ==============================
# TRATAMENTO DE AGÊNCIA E CONTA
# ==============================

# ==============================
# SPECS BANCÁRIOS
# Fonte: "Regras de Protocolação e Validação Bancária" (PDF oficial)
#
# Estrutura: código → BankSpec(ag_num, ag_dv, cc_num, cc_dv, nivel)
#
#   ag_num : dígitos esperados na agência (sem DV). None = variável (Nível 2).
#   ag_dv  : valor fixo do DV da agência (string "0","9"…) ou None (variável).
#            Quando ag_num e ag_dv são ambos fixos, a agência inteira é fixa
#            (ex: Inter Ag.0001-9 → ag_num="0001", ag_dv="9").
#   cc_num : dígitos esperados na conta (sem DC). None = variável.
#   cc_dv  : valor fixo do DC ou None (variável).
#   nivel  : 1 = regra específica | 2 = genérica | 3 = exceção
#
# Bancos com ag_num/ag_dv fixos (bancos digitais: 0001-D) são listados em
# AG_FIXA para facilitar a verificação de igualdade.
# ==============================

from dataclasses import dataclass
from typing import Optional

@dataclass
class BankSpec:
    ag_num  : Optional[int]    # qtd. dígitos do número da agência
    ag_dv   : Optional[str]    # DV fixo da agência ("0","9"…) ou None
    cc_num  : Optional[int]    # qtd. dígitos do número da conta
    cc_dv   : Optional[str]    # DV fixo da conta ou None
    nivel   : int              # 1, 2 ou 3
    # Quando True, o ag_num é um valor literal fixo (ex: "0001"), não um comprimento
    ag_num_fixo : bool = False

SPECS = {
    # ── Nível 1 ──────────────────────────────────────────────────────────────
    "001": BankSpec(4, None, 8, None, 1),          # BB          Ag:DDDD-D  CC:DDDDDDDD-D
    "033": BankSpec(4, "0",  8, None, 1),          # Santander   Ag:DDDD-0  CC:DDDDDDDD-D
    "041": BankSpec(4, None, 9, None, 1),          # Banrisul    Ag:DDDD-D  CC:DDDDDDDDD-D
    "070": BankSpec(4, "0",  6, None, 1),          # BRB         Ag:DDDD-0  CC:DDDDDD-D
    "077": BankSpec(4, "9",  7, None, 1, True),    # Inter       Ag:0001-9  CC:DDDDDDD-D
    "104": BankSpec(4, "0",  8, None, 1),          # Caixa       Ag:DDDD-0  CC:DDDDDDDD-D
    "133": BankSpec(4, "0",  7, None, 1),          # Cresol      Ag:DDDD-0  CC:DDDDDDD-D
    "136": BankSpec(4, "0",  7, None, 1),          # Unicred     Ag:DDDD-0  CC:DDDDDDD-D
    "197": BankSpec(4, "0",  8, None, 1, True),    # Stone       Ag:0001-0  CC:DDDDDDDD-D
    "208": BankSpec(4, "0",  8, None, 1),          # BTG         Ag:DDDD-0  CC:DDDDDDDD-D
    "212": BankSpec(4, "0",  8, None, 1, True),    # Original    Ag:0001-0  CC:DDDDDDDD-D
    "237": BankSpec(4, None, 7, None, 1),          # Bradesco    Ag:DDDD-D  CC:DDDDDDD-D
    "260": BankSpec(4, "0",  8, None, 1, True),    # Nubank      Ag:0001-0  CC:DDDDDDDD-D
    "290": BankSpec(4, "0",  8, None, 1, True),    # PagBank     Ag:0001-0  CC:DDDDDDDD-D
    "318": BankSpec(4, "0",  7, None, 1),          # BMG         Ag:DDDD-0  CC:DDDDDDD-D
    "323": BankSpec(4, "0",  8, None, 1, True),    # MercadoPago Ag:0001-0  CC:DDDDDDDD-D
    "336": BankSpec(4, "0",  6, None, 1, True),    # C6          Ag:0001-0  CC:DDDDDD-D
    "341": BankSpec(4, "0",  5, None, 1),          # Itaú        Ag:DDDD-0  CC:DDDDD-D
    "348": BankSpec(4, "0",  8, None, 1, True),    # XP          Ag:0001-0  CC:DDDDDDDD-D
    "364": BankSpec(4, "0",  8, None, 1, True),    # Efí         Ag:0001-0  CC:DDDDDDDD-D
    "380": BankSpec(4, "0",  8, None, 1, True),    # PicPay      Ag:0001-0  CC:DDDDDDDD-D
    "422": BankSpec(4, "0",  7, None, 1),          # Safra       Ag:DDDD-0  CC:DDDDDDD-D
    "623": BankSpec(4, "0",  8, None, 1),          # PAN         Ag:DDDD-0  CC:DDDDDDDD-D
    "748": BankSpec(4, "0",  7, None, 1),          # Sicredi     Ag:DDDD-0  CC:DDDDDDD-D
    "756": BankSpec(4, "0",  7, None, 1),          # Sicoob      Ag:DDDD-0  CC:DDDDDDD-D
    # ── Nível 2 ──────────────────────────────────────────────────────────────
    "004": BankSpec(None, None, None, None, 2),    # BNB
    "037": BankSpec(None, None, None, None, 2),    # Banpará
    "047": BankSpec(None, None, None, None, 2),    # Banese
    "084": BankSpec(None, None, None, None, 2),    # Uniprime
    "085": BankSpec(None, None, None, None, 2),    # Cecred
    "633": BankSpec(None, None, None, None, 2),    # Rendimento
    "637": BankSpec(None, None, None, None, 2),    # Sofisa
    "655": BankSpec(None, None, None, None, 2),    # Votorantim
    "707": BankSpec(None, None, None, None, 2),    # Daycoval
    "735": BankSpec(None, None, None, None, 2),    # Neon
    "739": BankSpec(None, None, None, None, 2),    # Cetelem
    "746": BankSpec(None, None, None, None, 2),    # Modal
}

# Valor literal fixo do número da agência para bancos digitais
AG_NUM_FIXO = {
    cod: "0001"
    for cod, sp in SPECS.items() if sp.ag_num_fixo
}

# ──────────────────────────────────────────────────────────────────────────────
# FILL LARANJA — marcador para células com inconsistência
# ──────────────────────────────────────────────────────────────────────────────
FILL_LARANJA = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# ──────────────────────────────────────────────────────────────────────────────
# RESULTADO DE CAMPO BANCÁRIO
# Cada campo (AG, DA, CC, DC) retorna um FieldResult.
#
#   valor    : string final padronizada (ou "" se ausente/inválida)
#   laranja  : True → célula deve ter fill laranja na planilha de saída
#   problemas: lista de strings descritivas para a coluna Exceções
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class FieldResult:
    valor    : str
    laranja  : bool
    problemas: list

    def ok(self): return not self.problemas


def _so_nums_hifens(s):
    """Remove tudo exceto dígitos e hífens. Retorna (limpo, tinha_outros)."""
    limpo = re.sub(r'[^\d\-]', '', s)
    tinha_outros = limpo != s
    return limpo, tinha_outros


def _so_nums_hifens_e_x(s):
    """Remove tudo exceto dígitos, hífens e a letra X (maiúscula/minúscula).
    Usado para DG e DC, onde X é um dígito verificador válido.
    Normaliza X para maiúsculo. Retorna (limpo, tinha_outros)."""
    limpo = re.sub(r'[^\dXx\-]', '', s).upper()
    tinha_outros = re.sub(r'[^\dXx\-]', '', s) != s
    return limpo, tinha_outros


def _so_nums(s):
    """Remove tudo exceto dígitos."""
    return re.sub(r'\D', '', s)


def _so_nums_e_x(s):
    """Remove tudo exceto dígitos e a letra X (maiúscula ou minúscula).
    Usado para extrair DV de campos que podem conter 'X' como dígito verificador.
    Normaliza X para maiúsculo."""
    return re.sub(r'[^\dXx]', '', s).upper()


def _vazio(raw):
    """True se a string não contém nenhum dígito."""
    return not any(c.isdigit() for c in (raw or ""))

def _vazio_dv(raw):
    """True se a string não contém nenhum dígito nem X (para DG/DA e DC,
    onde X é um dígito verificador válido)."""
    s = (raw or "").upper()
    return not any(c.isdigit() or c == 'X' for c in s)


def _split_hifen(s):
    """
    Divide em (antes, depois) no ÚLTIMO hífen.
    Retorna (None, None) se não houver hífen.
    """
    if '-' not in s:
        return None, None
    idx = s.rfind('-')
    return s[:idx], s[idx+1:]


def _dv_valido(s):
    """
    Retorna True se `s` é um dígito verificador válido para DG ou DC:
    exatamente 1 caractere numérico (0-9) OU a letra X (maiúscula ou minúscula),
    conforme especificado para os campos DG e DC.
    """
    if not s:
        return False
    return len(s) == 1 and (s.isdigit() or s.upper() == 'X')


# ──────────────────────────────────────────────────────────────────────────────
# TRATAMENTO DE AG (Agência)
# ──────────────────────────────────────────────────────────────────────────────

def tratar_ag(ag_raw, da_raw, spec):
    """
    Aplica o fluxo completo para AG e, quando há hífen em AG, também resolve DA.
    Retorna (res_ag: FieldResult, res_da_override: FieldResult|None).
    res_da_override != None significa que DA foi resolvido aqui (pular tratar_da).
    """
    ag_raw = str(ag_raw or "").strip()
    da_raw = str(da_raw or "").strip()

    # ── 5. Dados ausentes ────────────────────────────────────────────────────
    if _vazio(ag_raw):
        return (
            FieldResult("", True, ["AG: ausência de dados tratáveis"]),
            None
        )

    # Bancos com agência inteiramente fixa (número literal "0001")
    ag_num_fixo_val = AG_NUM_FIXO.get(spec and spec.ag_num_fixo and "xxx" or "")
    # Recuperar corretamente:
    ag_num_fixo_val = None
    if spec and spec.ag_num_fixo and spec.ag_num is not None:
        ag_num_fixo_val = "0001"   # todos os bancos digitais usam 0001

    # ── Agência com valor FIXO (número inteiro) ──────────────────────────────
    if ag_num_fixo_val is not None:
        nums_ag = _so_nums(ag_raw)
        # Comparar apenas os dígitos do número informado (ignorar hífen e DV)
        # ex: "0001-9" → nums antes do hífen = "0001"
        antes, depois = _split_hifen(nums_ag) if '-' in ag_raw else (nums_ag, None)
        num_ag_informado = _so_nums(antes) if antes is not None else _so_nums(ag_raw.split('-')[0])

        if num_ag_informado == ag_num_fixo_val:
            res_ag = FieldResult(ag_num_fixo_val, False, [])
        else:
            res_ag = FieldResult(
                ag_num_fixo_val, True,
                [f"AG: agência informada '{num_ag_informado}' difere do valor fixo "
                 f"'{ag_num_fixo_val}' — corrigido automaticamente"]
            )

        # DA também é fixo para esses bancos — retornar como override
        da_fixo = spec.ag_dv
        da_num_bruto = _so_nums_e_x(ag_raw.split('-')[-1]) if '-' in ag_raw else _so_nums_e_x(da_raw)
        if da_num_bruto == da_fixo:
            res_da = FieldResult(da_fixo, False, [])
        else:
            motivo = (f"DA: dígito informado '{da_num_bruto}' difere do valor fixo "
                      f"'{da_fixo}' — corrigido automaticamente")
            res_da = FieldResult(da_fixo, True, [motivo])
        return res_ag, res_da

    # ── Agência com comprimento variável (Nível 2) ───────────────────────────
    if spec is None or spec.ag_num is None:
        limpo, tinha_outros = _so_nums_hifens(ag_raw)
        problemas = []
        laranja = False
        if tinha_outros:
            problemas.append(f"AG: caracteres inválidos removidos de '{ag_raw}'")
            laranja = True

        if '-' in ag_raw:
            # Split no raw para preservar X no DV
            antes_h_raw, depois_h_raw = _split_hifen(ag_raw)
            num = _so_nums(re.sub(r'[^\d]', '', antes_h_raw))
            # Sem convenção: aceitar como está
            res_ag = FieldResult(num, laranja, problemas)
            # DA override: usar depois do hífen
            da_inline = _so_nums_e_x(depois_h_raw) if depois_h_raw else ""
            da_fornecido = _so_nums_e_x(da_raw)
            if da_inline and da_fornecido:
                if da_inline == da_fornecido:
                    res_da = FieldResult(da_inline, False, [])
                else:
                    res_da = FieldResult(
                        da_inline, True,
                        [f"DA: dígito no hífen de AG ('{da_inline}') difere do DA "
                         f"fornecido ('{da_fornecido}') — usado o do hífen"]
                    )
            elif da_inline:
                res_da = FieldResult(da_inline, False, [])
            elif da_fornecido:
                res_da = None   # deixar tratar_da cuidar
            else:
                res_da = FieldResult("", True, ["DA: ausência de dados tratáveis"])
            return res_ag, res_da
        else:
            num = _so_nums(limpo)
            return FieldResult(num, laranja, problemas), None

    # ── Agência com convenção específica (Nível 1) ───────────────────────────
    limpo, tinha_outros = _so_nums_hifens(ag_raw)
    problemas = []
    laranja = False
    if tinha_outros:
        problemas.append(f"AG: caracteres inválidos removidos de '{ag_raw}'")
        laranja = True

    ag_digits = spec.ag_num   # comprimento esperado do número da agência
    da_fixo   = spec.ag_dv    # DV fixo ou None

    res_da_override = None

    if '-' in ag_raw:
        # Fazer split no ag_raw original para preservar X no sufixo (DV)
        antes_h_raw, depois_h_raw = _split_hifen(ag_raw)
        num_ag = _so_nums(re.sub(r'[^\d]', '', antes_h_raw))
        da_hifen = _so_nums_e_x(depois_h_raw) if depois_h_raw else ""
        da_fornecido = _so_nums_e_x(da_raw)

        # Validar comprimento do número antes do hífen
        if len(num_ag) < ag_digits:
            num_ag = num_ag.zfill(ag_digits)
        elif len(num_ag) > ag_digits:
            problemas.append(
                f"AG: número da agência com {len(num_ag)} dígito(s) "
                f"(esperado {ag_digits}): '{num_ag}'"
            )
            laranja = True

        # Resolver DA usando os dois candidatos (hífen e coluna DA)
        da_hifen_valido   = _dv_valido(da_hifen)  if da_fixo is None else (da_hifen == da_fixo)
        da_fornec_valido  = _dv_valido(da_fornecido) if da_fixo is None else (da_fornecido == da_fixo)

        if da_hifen == da_fornecido or (not da_fornecido and da_hifen):
            # Iguais, ou só hífen disponível
            if da_hifen_valido:
                res_da_override = FieldResult(da_hifen, False, [])
            else:
                res_da_override = FieldResult(
                    da_hifen, True,
                    [f"DA: dígito '{da_hifen}' não segue a convenção "
                     + (f"(esperado '{da_fixo}')" if da_fixo else "(1 dígito esperado)")]
                )
        elif da_hifen_valido and not da_fornec_valido:
            res_da_override = FieldResult(
                da_hifen, True,
                [f"DA: usado dígito do hífen de AG ('{da_hifen}'); "
                 f"DA fornecido ('{da_fornecido}') não segue a convenção"]
            )
        elif da_fornec_valido and not da_hifen_valido:
            res_da_override = FieldResult(
                da_fornecido, True,
                [f"DA: usado DA fornecido ('{da_fornecido}'); "
                 f"dígito do hífen de AG ('{da_hifen}') não segue a convenção"]
            )
        elif da_hifen_valido and da_fornec_valido:
            # Ambos válidos mas diferentes → priorizar hífen (contíguo ao número)
            res_da_override = FieldResult(
                da_hifen, True,
                [f"DA: conflito — hífen de AG ('{da_hifen}') e DA fornecido "
                 f"('{da_fornecido}') diferem; ambos válidos — usado o do hífen"]
            )
        else:
            # Nenhum segue a convenção: usar DA fornecido, reportar
            dv_final = da_fornecido or da_hifen
            res_da_override = FieldResult(
                dv_final, True,
                [f"DA: nenhum dos dígitos disponíveis (hífen='{da_hifen}', "
                 f"fornecido='{da_fornecido}') segue a convenção — copiado como está"]
            )

        laranja = laranja or bool(problemas)
        return FieldResult(num_ag, laranja, problemas), res_da_override

    else:
        # Sem hífen em AG
        num_ag = _so_nums(limpo)
        if len(num_ag) < ag_digits:
            num_ag = num_ag.zfill(ag_digits)
        elif len(num_ag) > ag_digits:
            problemas.append(
                f"AG: número da agência com {len(num_ag)} dígito(s) "
                f"(esperado {ag_digits}): '{num_ag}'"
            )
            laranja = True
        return FieldResult(num_ag, laranja, problemas), None


# ──────────────────────────────────────────────────────────────────────────────
# TRATAMENTO DE DA (Dígito da Agência)
# Chamado apenas quando tratar_ag NÃO retornou res_da_override.
# ──────────────────────────────────────────────────────────────────────────────

def tratar_da(da_raw, ag_num_tratado, spec):
    """
    Retorna FieldResult para o DA.
    ag_num_tratado: número da agência já padronizado (para comparação com hífen).
    """
    da_raw = str(da_raw or "").strip()

    if _vazio_dv(da_raw):
        return FieldResult("", True, ["DA: ausência de dados tratáveis"])

    da_fixo = spec.ag_dv if spec else None

    # ── DV fixo ──────────────────────────────────────────────────────────────
    if da_fixo is not None:
        da_nums = _so_nums_e_x(da_raw)
        # Se houver hífen, o split já foi tratado em tratar_ag; aqui consideramos
        # só o valor puro da coluna DA.
        if '-' in da_raw:
            antes_h, depois_h = _split_hifen(da_raw)
            num_antes = _so_nums(antes_h)
            da_depois = _so_nums_e_x(depois_h) if depois_h else ""
            if num_antes == ag_num_tratado:
                dv_cand = da_depois
            else:
                dv_cand = da_depois
                problemas_extra = [f"DA: parte antes do hífen ('{num_antes}') "
                                   f"difere da AG tratada ('{ag_num_tratado}') — dado suspeito"]
            dv_cand_valido = (dv_cand == da_fixo)
            if dv_cand_valido:
                return FieldResult(dv_cand, False, getattr(locals(),'problemas_extra',[]))
            else:
                ps = getattr(locals(),'problemas_extra',[])
                ps.append(f"DA: dígito '{dv_cand}' difere do valor fixo '{da_fixo}' — corrigido")
                return FieldResult(da_fixo, True, ps)
        else:
            if da_nums == da_fixo:
                return FieldResult(da_fixo, False, [])
            else:
                return FieldResult(
                    da_fixo, True,
                    [f"DA: dígito informado '{da_nums}' difere do valor fixo "
                     f"'{da_fixo}' — corrigido automaticamente"]
                )

    # ── DV variável ───────────────────────────────────────────────────────────
    limpo, tinha_outros = _so_nums_hifens_e_x(re.sub(r'\.', '', da_raw))
    # X já está preservado; remover pontos já foi feito acima
    limpo_sem_pontos = limpo
    problemas = []
    laranja = False
    if tinha_outros:
        problemas.append(f"DA: caracteres inválidos removidos de '{da_raw}'")
        laranja = True

    if '-' in limpo_sem_pontos:
        antes_h, depois_h = _split_hifen(limpo_sem_pontos)
        num_antes = _so_nums(antes_h)
        da_depois = _so_nums_e_x(depois_h) if depois_h else ""

        if num_antes == ag_num_tratado or num_antes == ag_num_tratado.lstrip("0"):
            # Antes bate com AG
            if _dv_valido(da_depois):
                return FieldResult(da_depois.upper(), laranja, problemas)
            else:
                # Não segue convenção → últimos N dígitos/X (N=1)
                da_trunc = da_depois[-1].upper() if da_depois else ""
                problemas.append(
                    f"DA: dígito após hífen ('{da_depois}') não segue a convenção "
                    f"(1 dígito ou X esperado) — truncado para '{da_trunc}'"
                )
                return FieldResult(da_trunc, True, problemas)
        else:
            problemas.append(
                f"DA: parte antes do hífen ('{num_antes}') difere da AG "
                f"('{ag_num_tratado}') — dado suspeito"
            )
            if _dv_valido(da_depois):
                return FieldResult(da_depois.upper(), True, problemas)
            else:
                # Não segue convenção → últimos N dígitos/X (N=1)
                da_trunc = da_depois[-1].upper() if da_depois else ""
                problemas.append(
                    f"DA: dígito após hífen ('{da_depois}') não segue a convenção "
                    f"(1 dígito ou X esperado) — truncado para '{da_trunc}'"
                )
                return FieldResult(da_trunc, True, problemas)
    else:
        # Para DG/DA: aceitar dígitos 0-9 ou X
        # Recalcular preservando X que _so_nums_hifens descartou
        da_bruto = _so_nums_e_x(re.sub(r'\.', '', da_raw))
        if _dv_valido(da_bruto):
            return FieldResult(da_bruto, laranja, problemas)
        else:
            # Mais de 1 char → manter apenas o último (começando do final)
            da_truncado = da_bruto[-1] if da_bruto else ""
            problemas.append(
                f"DA: valor '{da_bruto}' não segue a convenção "
                f"(1 dígito ou X esperado) — truncado para '{da_truncado}'"
            )
            return FieldResult(da_truncado, True, problemas)


# ──────────────────────────────────────────────────────────────────────────────
# TRATAMENTO DE CC (Conta Corrente) e DC (Dígito da Conta)
# ──────────────────────────────────────────────────────────────────────────────

def tratar_cc_dc(cc_raw, dc_raw, spec):
    """
    Aplica o fluxo completo para CC e DC.
    Retorna (res_cc: FieldResult, res_dc: FieldResult).
    """
    cc_raw = str(cc_raw or "").strip()
    dc_raw = str(dc_raw or "").strip()

    cc_digits = spec.cc_num if spec else None
    dc_fixo   = spec.cc_dv  if spec else None

    # ── 5. Dados ausentes ────────────────────────────────────────────────────
    cc_vazio = _vazio(cc_raw)
    dc_vazio = _vazio_dv(dc_raw)   # DC pode ser 'X' como dígito verificador

    if cc_vazio:
        res_cc = FieldResult("", True, ["CC: ausência de dados tratáveis"])
        if dc_vazio:
            res_dc = FieldResult("", True, ["DC: ausência de dados tratáveis"])
        else:
            res_dc = _tratar_dc_puro(dc_raw, "", spec)
        return res_cc, res_dc

    # Remover pontos (Banrisul) — excluir pontos no pré-processamento de CC
    cc_sem_pontos = re.sub(r'\.', '', cc_raw)

    limpo_cc, tinha_outros_cc = _so_nums_hifens(cc_sem_pontos)
    problemas_cc = []
    laranja_cc   = False
    if tinha_outros_cc:
        problemas_cc.append(f"CC: caracteres inválidos removidos de '{cc_raw}'")
        laranja_cc = True

    # ── Nível 2: sem convenção específica ────────────────────────────────────
    if cc_digits is None:
        if '-' in cc_sem_pontos:
            antes_h_raw, depois_h_raw = _split_hifen(cc_sem_pontos)
            num_cc = _so_nums(antes_h_raw)
            dc_hifen = _so_nums_e_x(depois_h_raw) if depois_h_raw else ""
            res_cc = FieldResult(num_cc, laranja_cc, problemas_cc)
            res_dc = _resolver_dc_com_hifen(dc_hifen, dc_raw, num_cc, spec)
        else:
            num_cc = _so_nums(limpo_cc)
            res_cc = FieldResult(num_cc, laranja_cc, problemas_cc)
            if dc_vazio:
                res_dc = FieldResult("", True, ["DC: ausência de dados tratáveis"])
            else:
                res_dc = _tratar_dc_puro(dc_raw, num_cc, spec)
        return res_cc, res_dc

    # ── Nível 1: convenção específica ─────────────────────────────────────────
    if '-' in cc_sem_pontos:
        antes_h_raw, depois_h_raw = _split_hifen(cc_sem_pontos)
        num_cc   = _so_nums(antes_h_raw)
        dc_hifen = _so_nums_e_x(depois_h_raw) if depois_h_raw else ""

        # Validar comprimento do número antes do hífen
        if len(num_cc) < cc_digits:
            num_cc = num_cc.zfill(cc_digits)
        elif len(num_cc) > cc_digits:
            problemas_cc.append(
                f"CC: número da conta com {len(num_cc)} dígito(s) "
                f"(esperado {cc_digits}): '{num_cc}'"
            )
            laranja_cc = True

        res_cc = FieldResult(num_cc, laranja_cc, problemas_cc)
        res_dc = _resolver_dc_com_hifen(dc_hifen, dc_raw, num_cc, spec)
        return res_cc, res_dc

    else:
        # Sem hífen na CC
        num_cc = _so_nums(limpo_cc)

        if len(num_cc) == cc_digits:
            # Comprimento exato
            res_cc = FieldResult(num_cc, laranja_cc, problemas_cc)
            if dc_vazio:
                res_dc = FieldResult("", True, ["DC: ausência de dados tratáveis"])
            else:
                res_dc = _tratar_dc_puro(dc_raw, num_cc, spec)
            return res_cc, res_dc

        elif len(num_cc) < cc_digits:
            # CC curta — tentar emenda com DC
            dc_nums = _so_nums_e_x(re.sub(r'\.', '', dc_raw))
            total = num_cc + dc_nums
            if len(total) == cc_digits + 1:
                # Emenda possível: dividir conforme convenção
                cc_novo = total[:cc_digits]
                dc_novo = total[cc_digits:]
                problemas_cc.append(
                    f"CC: conta curta + DC emendados — CC='{cc_novo}', DC='{dc_novo}' — confirmar"
                )
                res_cc = FieldResult(cc_novo, True, problemas_cc)
                res_dc = FieldResult(dc_novo, True,
                                     ["DC: dígito obtido por emenda com CC — confirmar"])
            else:
                # Emenda impossível — zfill
                num_cc_pad = num_cc.zfill(cc_digits)
                problemas_cc.append(
                    f"CC: conta curta ({len(num_cc)}d, esperado {cc_digits}d) — "
                    f"zeros adicionados à esquerda"
                )
                res_cc = FieldResult(num_cc_pad, True, problemas_cc)
                if dc_vazio:
                    res_dc = FieldResult("", True, ["DC: ausência de dados tratáveis"])
                else:
                    res_dc = _tratar_dc_puro(dc_raw, num_cc_pad, spec)
            return res_cc, res_dc

        else:
            # CC longa — verificar se o excedente pode ser o DC
            excedente_len = len(num_cc) - cc_digits
            dc_nums = _so_nums_e_x(re.sub(r'\.', '', dc_raw)) if not dc_vazio else ""
            dc_fixo_local = spec.cc_dv if spec else None
            dc_esperado_len = 1   # convenção: 1 dígito para DC

            if excedente_len == dc_esperado_len:
                excedente = num_cc[-excedente_len:]
                cc_sem_exc = num_cc[:-excedente_len]
                if excedente == dc_nums:
                    # Excedente bate com DC informado → tratar como CC+DC emendados
                    problemas_cc.append(
                        f"CC: conta com {len(num_cc)}d contém DC embutido — "
                        f"CC='{cc_sem_exc}', DC='{excedente}' — confirmar"
                    )
                    res_cc = FieldResult(cc_sem_exc, True, problemas_cc)
                    res_dc = FieldResult(excedente, True,
                                         ["DC: dígito extraído do excedente de CC — confirmar"])
                else:
                    # Excedente difere do DC informado → retirar excedente e reportar
                    problemas_cc.append(
                        f"CC: conta com {len(num_cc)}d (esperado {cc_digits}d) — "
                        f"excedente '{excedente}' removido (DC informado='{dc_nums}')"
                    )
                    res_cc = FieldResult(cc_sem_exc, True, problemas_cc)
                    res_dc = _tratar_dc_puro(dc_raw, cc_sem_exc, spec) if not dc_vazio else FieldResult("", True, ["DC: ausência de dados tratáveis"])
            else:
                # Excedente com comprimento diferente do esperado → remover e reportar
                cc_sem_exc = num_cc[:cc_digits]
                problemas_cc.append(
                    f"CC: conta com {len(num_cc)}d (esperado {cc_digits}d) — "
                    f"truncado para '{cc_sem_exc}'"
                )
                res_cc = FieldResult(cc_sem_exc, True, problemas_cc)
                res_dc = _tratar_dc_puro(dc_raw, cc_sem_exc, spec) if not dc_vazio else FieldResult("", True, ["DC: ausência de dados tratáveis"])
            return res_cc, res_dc


def _resolver_dc_com_hifen(dc_hifen, dc_raw, cc_num_tratado, spec):
    """
    Resolve DC quando a CC tinha hífen.
    dc_hifen: string após o hífen em CC (só dígitos).
    dc_raw: valor bruto da coluna DC.
    """
    dc_fixo  = spec.cc_dv if spec else None
    dc_fornecido = _so_nums_e_x(re.sub(r'\.', '', str(dc_raw or "")))

    dc_hifen_valido  = _dv_valido(dc_hifen)  if dc_fixo is None else (dc_hifen == dc_fixo)
    dc_fornec_valido = _dv_valido(dc_fornecido) if dc_fixo is None else (dc_fornecido == dc_fixo)

    # Helper: truncar para últimos N dígitos quando não segue convenção
    def _dc_truncar(val, dc_fixo_local):
        if dc_fixo_local is not None:
            return dc_fixo_local   # fixo: já corrigido antes de chegar aqui
        return val[-1] if val else ""

    if dc_hifen == dc_fornecido or (dc_hifen and not dc_fornecido):
        if dc_hifen_valido:
            return FieldResult(dc_hifen, False, [])
        else:
            dc_t = _dc_truncar(dc_hifen, dc_fixo)
            return FieldResult(
                dc_t, True,
                [f"DC: dígito '{dc_hifen}' não segue a convenção"
                 + (f" (esperado '{dc_fixo}')" if dc_fixo else "")
                 + (f" — truncado para '{dc_t}'" if dc_t != dc_hifen else "")]
            )
    elif dc_hifen_valido and not dc_fornec_valido:
        return FieldResult(
            dc_hifen, True,
            [f"DC: usado dígito do hífen de CC ('{dc_hifen}'); "
             f"DC fornecido ('{dc_fornecido}') não segue a convenção"]
        )
    elif dc_fornec_valido and not dc_hifen_valido:
        return FieldResult(
            dc_fornecido, True,
            [f"DC: usado DC fornecido ('{dc_fornecido}'); "
             f"dígito do hífen de CC ('{dc_hifen}') não segue a convenção"]
        )
    elif dc_hifen_valido and dc_fornec_valido:
        # Ambos válidos mas diferentes → priorizar hífen (contíguo ao número)
        return FieldResult(
            dc_hifen, True,
            [f"DC: conflito — hífen de CC ('{dc_hifen}') e DC fornecido "
             f"('{dc_fornecido}') diferem; ambos válidos — usado o do hífen"]
        )
    else:
        # Nenhum segue a convenção → últimos N dígitos de cada candidato disponível
        cand = dc_hifen or dc_fornecido
        dc_t = _dc_truncar(cand, dc_fixo)
        return FieldResult(
            dc_t, True,
            [f"DC: nenhum dos dígitos disponíveis (hífen='{dc_hifen}', "
             f"fornecido='{dc_fornecido}') segue a convenção — truncado para '{dc_t}'"]
        )


def _tratar_dc_puro(dc_raw, cc_num_tratado, spec):
    """
    Trata a coluna DC isoladamente (sem hífen na CC).
    Equivale ao fluxo '4. Para DC' completo.
    """
    dc_raw = str(dc_raw or "").strip()
    if _vazio_dv(dc_raw):
        return FieldResult("", True, ["DC: ausência de dados tratáveis"])

    dc_fixo = spec.cc_dv if spec else None

    # DV fixo
    if dc_fixo is not None:
        dc_nums = _so_nums_e_x(re.sub(r'\.', '', dc_raw))
        if '-' in dc_raw:
            antes_h, depois_h = _split_hifen(re.sub(r'\.', '', dc_raw))
            num_antes = _so_nums(antes_h)
            dc_depois = _so_nums_e_x(depois_h) if depois_h else ""
            ps = []
            if num_antes != cc_num_tratado and num_antes != cc_num_tratado.lstrip("0"):
                ps.append(f"DC: parte antes do hífen ('{num_antes}') difere da CC "
                          f"('{cc_num_tratado}') — dado suspeito")
            if dc_depois == dc_fixo:
                return FieldResult(dc_fixo, bool(ps), ps)
            else:
                ps.append(f"DC: dígito '{dc_depois}' difere do valor fixo '{dc_fixo}' — corrigido")
                return FieldResult(dc_fixo, True, ps)
        else:
            if dc_nums == dc_fixo:
                return FieldResult(dc_fixo, False, [])
            else:
                return FieldResult(
                    dc_fixo, True,
                    [f"DC: dígito informado '{dc_nums}' difere do valor fixo "
                     f"'{dc_fixo}' — corrigido automaticamente"]
                )

    # DV variável
    limpo, tinha_outros = _so_nums_hifens_e_x(re.sub(r'\.', '', dc_raw))
    problemas = []
    laranja   = False
    if tinha_outros:
        problemas.append(f"DC: caracteres inválidos removidos de '{dc_raw}'")
        laranja = True

    if '-' in limpo:
        antes_h, depois_h = _split_hifen(limpo)
        num_antes = _so_nums(antes_h)
        dc_depois = _so_nums(depois_h) if depois_h else ""
        if num_antes != cc_num_tratado and num_antes != cc_num_tratado.lstrip("0"):
            problemas.append(f"DC: parte antes do hífen ('{num_antes}') difere da CC "
                             f"('{cc_num_tratado}') — dado suspeito")
            laranja = True
        if _dv_valido(dc_depois):
            return FieldResult(dc_depois.upper(), laranja, problemas)
        else:
            # Não segue convenção → últimos N dígitos/X (N=1)
            dc_t = dc_depois[-1].upper() if dc_depois else ""
            problemas.append(
                f"DC: dígito após hífen ('{dc_depois}') não segue a convenção "
                f"(1 dígito ou X esperado) — truncado para '{dc_t}'"
            )
            return FieldResult(dc_t, True, problemas)
    else:
        # Para DC: aceitar dígitos 0-9 ou X
        # Recalcular preservando X que _so_nums_hifens descartou
        dc_bruto = _so_nums_e_x(re.sub(r'\.', '', dc_raw))
        if _dv_valido(dc_bruto):
            return FieldResult(dc_bruto, laranja, problemas)
        else:
            # Não segue convenção → manter apenas o último char
            dc_t = dc_bruto[-1] if dc_bruto else ""
            problemas.append(
                f"DC: valor '{dc_bruto}' não segue a convenção "
                f"(1 dígito ou X esperado) — truncado para '{dc_t}'"
            )
            return FieldResult(dc_t, True, problemas)


# ──────────────────────────────────────────────────────────────────────────────
# PONTO DE ENTRADA: tratar_dados_bancarios
# ──────────────────────────────────────────────────────────────────────────────

def tratar_dados_bancarios(ag_raw, da_raw, cc_raw, dc_raw, codigo_banco):
    """
    Orquestra o tratamento completo dos 4 campos bancários conforme o fluxo
    especificado, usando as convenções do PDF oficial.

    Retorna:
        ag   : str   — número da agência padronizado
        da   : str   — dígito da agência padronizado
        cc   : str   — número da conta padronizado
        dc   : str   — dígito da conta padronizado
        status : int — 0=sem dados | 1=N1 | 2=N2 | 3=exceção
        excecao: str — descrição(ões) concatenadas (vazio se ok)
        laranja: dict{campo: bool} — quais campos ficam laranja
                 campos: "ag","da","cc","dc"
    """
    spec = SPECS.get(codigo_banco)

    # Sem regra cadastrada → Nível 3
    if spec is None and codigo_banco:
        spec_fake = BankSpec(None, None, None, None, 3)
        msg = (f"Banco '{codigo_banco}' sem regra de validação cadastrada "
               "— conferência manual necessária")
        # Tratar como Nível 2 nos valores mas marcar tudo como exceção
        res_ag, res_da_ovr = tratar_ag(ag_raw, da_raw, None)
        if res_da_ovr is not None:
            res_da = res_da_ovr
        else:
            res_da = tratar_da(da_raw, res_ag.valor, None)
        res_cc, res_dc = tratar_cc_dc(cc_raw, dc_raw, None)

        todos_problemas = ([msg]
                           + res_ag.problemas + res_da.problemas
                           + res_cc.problemas + res_dc.problemas)
        return (
            res_ag.valor, res_da.valor, res_cc.valor, res_dc.valor,
            3,
            "; ".join(todos_problemas),
            {"ag": True, "da": True, "cc": True, "dc": True},
        )

    # ── AG e DA ──────────────────────────────────────────────────────────────
    res_ag, res_da_ovr = tratar_ag(ag_raw, da_raw, spec)
    if res_da_ovr is not None:
        res_da = res_da_ovr
    else:
        res_da = tratar_da(da_raw, res_ag.valor, spec)

    # ── CC e DC ──────────────────────────────────────────────────────────────
    res_cc, res_dc = tratar_cc_dc(cc_raw, dc_raw, spec)

    # ── Status ───────────────────────────────────────────────────────────────
    todos_problemas = (res_ag.problemas + res_da.problemas
                       + res_cc.problemas + res_dc.problemas)

    nivel = spec.nivel if spec else 3
    if not todos_problemas:
        status = nivel
    else:
        # Verificar se há problemas graves (que rebaixam para 3)
        graves = [p for p in todos_problemas
                  if not any(x in p for x in ("zeros adicionados", "confirmar",
                                               "removidos", "ausência"))]
        status = 3 if graves else nivel

    # Sem nenhum dado útil
    todos_vazios = (not res_ag.valor and not res_da.valor
                    and not res_cc.valor and not res_dc.valor)
    if todos_vazios:
        status = 0

    return (
        res_ag.valor,
        res_da.valor,
        res_cc.valor,
        res_dc.valor,
        status,
        "; ".join(todos_problemas),
        {
            "ag": res_ag.laranja,
            "da": res_da.laranja,
            "cc": res_cc.laranja,
            "dc": res_dc.laranja,
        },
    )


# ==============================
# TRATAMENTO DE E-MAIL# ==============================
# TRATAMENTO DE E-MAIL
# ==============================

def padronizar_email(valor):
    if not valor or normalizar_texto(valor) in ("", "nan"):
        return ("", "")

    email = str(valor).strip().lower()

    # Domínios não-corporativos conhecidos
    DOMINIOS_PESSOAIS = {"gmail", "hotmail", "outlook", "yahoo", "bol", "uol", "live", "icloud"}

    # Remove espaços internos
    email = re.sub(r'\s+', '', email)

    # Verifica formato básico
    match = re.match(r'^([^@]+)@([^@]+)$', email)
    if not match:
        return (email, f"E-mail com formato inválido: '{email}'")

    usuario, dominio = match.group(1), match.group(2)

    # Verifica TLD: precisa de pelo menos um ponto
    partes_dominio = dominio.split('.')
    if len(partes_dominio) < 2 or any(p == '' for p in partes_dominio):
        # Tenta corrigir se for um domínio pessoal sem TLD
        dominio_base = partes_dominio[0]
        if dominio_base in DOMINIOS_PESSOAIS:
            email_corrigido = f"{usuario}@{dominio_base}.com"
            return (email_corrigido, f"E-mail incompleto corrigido: '{valor}' → '{email_corrigido}'")
        return (email, f"Domínio de e-mail possivelmente incompleto: '{dominio}'")

    # Avisa sobre domínio pessoal (mas não rejeita)
    dominio_base = partes_dominio[0]
    if dominio_base in DOMINIOS_PESSOAIS:
        return (email, f"E-mail não institucional (pessoal): '{email}'")

    return (email, "")


# ==============================
# TRATAMENTO DE CPF
# ==============================

def padronizar_cpf(valor):
    if not valor or normalizar_texto(valor) in ("", "nan"):
        return ("", "")
    cpf = re.sub(r'\D', '', str(valor))
    if len(cpf) != 11:
        return (cpf, f"CPF com número de dígitos incorreto: '{valor}'")
    return (cpf, "")


# ==============================
# TRATAMENTO DE DATA
# ==============================

def padronizar_data(valor):
    if not valor or normalizar_texto(valor) in ("", "nan"):
        return (None, "")
    if isinstance(valor, (datetime, date)):
        return (valor if isinstance(valor, datetime) else datetime.combine(valor, datetime.min.time()), "")
    try:
        return (pd.to_datetime(str(valor), dayfirst=True), "")
    except Exception:
        return (None, f"Data não reconhecida: '{valor}'")


def calcular_maior_idade(data_nascimento):
    if data_nascimento is None:
        return None
    hoje = datetime.today()
    anos = hoje.year - data_nascimento.year - (
        (hoje.month, hoje.day) < (data_nascimento.month, data_nascimento.day)
    )
    return anos >= 18


# ==============================
# TRATAMENTO DE CEP
# ==============================

def padronizar_cep(valor):
    if not valor or normalizar_texto(valor) in ("", "nan"):
        return ("", "")
    cep = re.sub(r'\D', '', str(valor))
    if len(cep) != 8:
        return (cep, f"CEP com número de dígitos incorreto: '{valor}'")
    return (cep, "")


# ==============================
# CORRESPONDÊNCIA DE NOMES
# ==============================

def nome_esta_contido(nome_aluno, nome_formulario):
    """
    Verifica se o nome do aluno (planilha alunos.xlsx) está contido
    no nome do formulário (dados_bancarios.xlsx), após normalização.
    A proposta exige que o primeiro esteja contido no segundo.
    """
    a = normalizar_nome(nome_aluno)
    b = normalizar_nome(nome_formulario)
    if not a or not b:
        return False
    # Tokeniza e verifica se todos os tokens do nome do aluno aparecem no formulário
    tokens_a = set(a.split())
    tokens_b = set(b.split())
    return tokens_a.issubset(tokens_b)


# ==============================
# LEITURA DAS PLANILHAS
# ==============================

def ler_alunos(path):
    df = ler_excel(path)
    df = normalizar_colunas(df)
    return df


def ler_dados_bancarios(path):
    df = ler_excel(path)
    df = normalizar_colunas(df)
    return df


def _todas_cols(df, keywords, excluir=None):
    """Retorna todas as colunas que contêm todas as keywords."""
    resultado = []
    for col in df.columns:
        if all(kw in col for kw in keywords):
            if excluir and any(ek in col for ek in excluir):
                continue
            resultado.append(col)
    return resultado


def detectar_conjuntos_bancarios(df):
    """
    Detecta os conjuntos de colunas bancárias no formulário.

    Estratégia posicional direta:
    - Localiza todas as colunas com "corrente" (número de conta corrente)
      como âncora, pois esse termo é inequívoco e aparece uma vez por conjunto.
    - Para cada coluna-conta encontrada, os campos do conjunto são as 3
      colunas imediatamente anteriores (banco, agência, dígito agência)
      e a coluna seguinte (dígito conta).

    Essa abordagem é insensível ao conteúdo dos títulos e funciona mesmo
    quando os cabeçalhos são longos, ambíguos ou em português com variações.

    Sobre a coluna "INDICAR O DÍGITO DA AGÊNCIA" (col[18] neste formulário):
    Apesar do título mencionar "agência", na prática essa coluna guarda o
    dígito da CONTA para usuários de bancos digitais no set1 — e está
    posicionalmente APÓS a conta (col[17]), confirmando ser dig_cc do set1.
    Como não existe coluna de dígito da agência separada no set1, dig_ag
    fica None para esse conjunto.
    """
    colunas = list(df.columns)

    # Encontra todas as colunas de conta corrente (âncora do conjunto)
    ancoras_conta = []
    for i, col in enumerate(colunas):
        if "corrente" in col and "responsavel" not in col and "digito" not in col:
            ancoras_conta.append(i)

    conjuntos = []
    usadas = set()

    for i_conta in ancoras_conta:
        if i_conta in usadas:
            continue

        col_conta = colunas[i_conta]

        # As 3 colunas antes da conta: banco, agencia, dig_ag
        # (podem não existir se a conta estiver no início)
        def col_antes(offset):
            idx = i_conta - offset
            if idx >= 0 and idx not in usadas:
                return idx, colunas[idx]
            return None, None

        i_dag, col_dag   = col_antes(1)   # 1 antes da conta = dígito agência
        i_ag,  col_ag    = col_antes(2)   # 2 antes = agência
        i_bk,  col_banco = col_antes(3)   # 3 antes = banco

        # Coluna de dígito da conta: a SEGUINTE à conta
        i_dcc  = i_conta + 1
        col_dcc = colunas[i_dcc] if i_dcc < len(colunas) else None
        # dig_cc não pode ser uma coluna sem "digito" (não pode ser outra conta)
        if col_dcc and "digito" not in col_dcc and "corrente" in col_dcc:
            col_dcc = None
            i_dcc   = None

        # Marca como usadas para não reutilizar em outro conjunto
        for idx in [i_bk, i_ag, i_dag, i_conta, i_dcc]:
            if idx is not None:
                usadas.add(idx)

        conjuntos.append({
            "banco":   col_banco,
            "agencia": col_ag,
            "dig_ag":  col_dag,
            "conta":   col_conta,
            "dig_cc":  col_dcc,
        })

    if not conjuntos:
        conjuntos = [{"banco": None, "agencia": None, "dig_ag": None,
                      "conta": None, "dig_cc": None}]

    return conjuntos


def _valor_nao_vazio(row, col):
    """Retorna True se a célula da coluna col na linha row tem conteúdo real."""
    if col is None:
        return False
    v = row.get(col, None)
    if v is None:
        return False
    return str(v).strip().lower() not in ("", "nan", "none")


def _escolher_conjunto(row, conjuntos):
    """
    Para uma linha de dados, escolhe o conjunto bancário que tem mais campos
    preenchidos. Em caso de empate, prefere o de índice mais alto (conjuntos
    mais recentes tendem a ser versões corrigidas do formulário).
    """
    melhor_idx = 0
    melhor_score = -1
    for i, c in enumerate(conjuntos):
        score = sum(_valor_nao_vazio(row, c[k])
                    for k in ("banco", "agencia", "dig_ag", "conta", "dig_cc"))
        if score >= melhor_score:
            melhor_score = score
            melhor_idx = i
    return conjuntos[melhor_idx]


# ==============================
# NÚCLEO DO ALGORITMO
# ==============================

def processar(path_alunos, path_bancarios, path_template=None):
    """
    Executa o processamento completo.
    Retorna (df_saida, logs_erros, caminho_template_usado).
    """
    logs = []

    def log(linha, linha_planilha=None):
        prefixo = f"[Linha {linha_planilha}] " if linha_planilha else ""
        logs.append(prefixo + linha)

    # --- Leitura ---
    df_alunos = ler_alunos(path_alunos)
    df_banco = ler_dados_bancarios(path_bancarios)

    # --- Detectar colunas em alunos.xlsx ---
    col_turma   = encontrar_coluna_por_keywords(df_alunos, ["turma"])
    col_nome_al = encontrar_coluna_por_keywords(df_alunos, ["nome"])
    col_email   = encontrar_coluna_por_keywords(df_alunos, ["email"])
    col_cpf_al  = encontrar_coluna_por_keywords(df_alunos, ["cpf"])
    col_rg      = encontrar_coluna_por_keywords(df_alunos, ["rg"], obrigatoria=False)
    col_nasc_al = encontrar_coluna_por_keywords(df_alunos, ["nasc"], obrigatoria=False) or \
                  encontrar_coluna_por_keywords(df_alunos, ["data"], obrigatoria=False)
    col_end     = encontrar_coluna_por_keywords(df_alunos, ["endereco"], obrigatoria=False) or \
                  encontrar_coluna_por_keywords(df_alunos, ["logradouro"], obrigatoria=False)
    col_bairro  = encontrar_coluna_por_keywords(df_alunos, ["bairro"], obrigatoria=False)
    col_cidade  = encontrar_coluna_por_keywords(df_alunos, ["cidade"], obrigatoria=False)
    col_estado  = encontrar_coluna_por_keywords(df_alunos, ["estado"], obrigatoria=False) or \
                  encontrar_coluna_por_keywords(df_alunos, ["uf"], obrigatoria=False)
    col_cep     = encontrar_coluna_por_keywords(df_alunos, ["cep"], obrigatoria=False)

    # --- Detectar colunas em dados_bancarios.xlsx ---
    col_nome_bk  = encontrar_coluna_por_keywords(df_banco, ["nome"], excluir_keywords=["responsavel"])
    col_cpf_bk   = encontrar_coluna_por_keywords(df_banco, ["cpf"], obrigatoria=False)
    col_nasc_bk  = encontrar_coluna_por_keywords(df_banco, ["nasc"], obrigatoria=False) or \
                   encontrar_coluna_por_keywords(df_banco, ["data", "nascimento"], obrigatoria=False)
    conjuntos_bancarios = detectar_conjuntos_bancarios(df_banco)

    # Responsável legal
    col_nome_resp   = encontrar_coluna_por_keywords(df_banco, ["responsavel", "nome"], obrigatoria=False)
    col_cpf_resp    = encontrar_coluna_por_keywords(df_banco, ["responsavel", "cpf"], obrigatoria=False)
    col_email_resp  = (encontrar_coluna_por_keywords(df_banco, ["responsavel", "email"],   obrigatoria=False) or
                       encontrar_coluna_por_keywords(df_banco, ["email",       "responsavel"], obrigatoria=False) or
                       encontrar_coluna_por_keywords(df_banco, ["responsavel", "e-mail"],      obrigatoria=False) or
                       encontrar_coluna_por_keywords(df_banco, ["e-mail",      "responsavel"], obrigatoria=False))
    col_cont_resp   = encontrar_coluna_por_keywords(df_banco, ["responsavel", "contato"], obrigatoria=False) or \
                      encontrar_coluna_por_keywords(df_banco, ["celular", "responsavel"], obrigatoria=False)

    # Ordenar turmas alfabeticamente, alunos alfabeticamente dentro de cada turma
    df_alunos["_turma_norm"] = df_alunos[col_turma].apply(normalizar_texto)
    df_alunos["_nome_norm"]  = df_alunos[col_nome_al].apply(normalizar_nome)
    df_alunos = df_alunos.sort_values(["_turma_norm", "_nome_norm"]).reset_index(drop=True)

    # Índice de busca no formulário (nome normalizado → linha(s))
    df_banco["_nome_norm"] = df_banco[col_nome_bk].apply(normalizar_nome)

    linhas_saida = []
    id_por_turma = {}

    for idx_al, row_al in df_alunos.iterrows():
        linha_planilha = len(linhas_saida) + 2  # +2: cabeçalho + 1-index

        nome_al = str(row_al[col_nome_al]).strip() if pd.notna(row_al[col_nome_al]) else ""
        turma   = str(row_al[col_turma]).strip() if pd.notna(row_al[col_turma]) else ""

        # --- ID por turma ---
        if turma not in id_por_turma:
            id_por_turma[turma] = 0
        id_por_turma[turma] += 1
        id_aluno = id_por_turma[turma]

        # --- Correspondência de nome ---
        matches = df_banco[df_banco["_nome_norm"].apply(
            lambda n: nome_esta_contido(nome_al, n)
        )]

        if len(matches) == 0:
            log(f"Sem correspondência no formulário para aluno: '{nome_al}' (turma: {turma})", linha_planilha)
            row_bk = None
        elif len(matches) > 1:
            log(f"Múltiplas correspondências no formulário para aluno: '{nome_al}' — usando primeira ocorrência", linha_planilha)
            row_bk = matches.iloc[0]
        else:
            row_bk = matches.iloc[0]

        # --- E-mail ---
        email_raw = str(row_al[col_email]).strip() if col_email and pd.notna(row_al.get(col_email)) else ""
        email, alerta_email = padronizar_email(email_raw)
        if alerta_email:
            log(alerta_email, linha_planilha)

        # --- CPF ---
        cpf_al_raw = str(row_al[col_cpf_al]).strip() if pd.notna(row_al.get(col_cpf_al)) else ""
        cpf, alerta_cpf = padronizar_cpf(cpf_al_raw)
        if alerta_cpf:
            log(alerta_cpf, linha_planilha)
        if row_bk is not None and col_cpf_bk:
            cpf_bk_raw = str(row_bk.get(col_cpf_bk, "")).strip()
            cpf_bk, _ = padronizar_cpf(cpf_bk_raw)
            if cpf and cpf_bk and cpf != cpf_bk:
                log(f"CPF divergente para '{nome_al}': alunos='{cpf}' / formulário='{cpf_bk}'", linha_planilha)

        # --- RG ---
        rg = str(row_al[col_rg]).strip() if col_rg and pd.notna(row_al.get(col_rg)) else ""

        # --- Data de Nascimento ---
        nasc_raw = str(row_al[col_nasc_al]).strip() if col_nasc_al and pd.notna(row_al.get(col_nasc_al)) else ""
        data_nasc, alerta_nasc = padronizar_data(nasc_raw)
        if alerta_nasc:
            log(alerta_nasc, linha_planilha)
        if row_bk is not None and col_nasc_bk:
            nasc_bk_raw = str(row_bk.get(col_nasc_bk, "")).strip()
            data_nasc_bk, _ = padronizar_data(nasc_bk_raw)
            if data_nasc and data_nasc_bk and data_nasc.date() != data_nasc_bk.date():
                log(f"Data de nascimento divergente para '{nome_al}': alunos='{data_nasc.date()}' / formulário='{data_nasc_bk.date()}'", linha_planilha)

        maior = calcular_maior_idade(data_nasc)

        # --- Endereço ---
        partes_end = []
        for col_e in [col_end, col_bairro, col_cidade, col_estado]:
            if col_e and pd.notna(row_al.get(col_e)):
                v = str(row_al[col_e]).strip()
                if v and v.lower() not in ("nan", ""):
                    partes_end.append(v)
        endereco_completo = " - ".join(partes_end)
        # Remove hífens duplicados que possam surgir se algum campo já contiver " - "
        endereco_completo = re.sub(r'(\s*-\s*){2,}', ' - ', endereco_completo).strip(" -")

        # --- CEP ---
        cep_raw = str(row_al[col_cep]).strip() if col_cep and pd.notna(row_al.get(col_cep)) else ""
        cep, alerta_cep = padronizar_cep(cep_raw)
        if alerta_cep:
            log(alerta_cep, linha_planilha)

        # --- Responsável legal ---
        nome_resp = cpf_resp = email_resp = cont_resp = ""
        if row_bk is not None:
            if maior is False:  # menor de idade
                nome_resp  = str(row_bk.get(col_nome_resp, "")).strip() if col_nome_resp else ""
                cpf_resp   = str(row_bk.get(col_cpf_resp, "")).strip() if col_cpf_resp else ""
                email_resp = str(row_bk.get(col_email_resp, "")).strip() if col_email_resp else ""
                cont_resp  = str(row_bk.get(col_cont_resp, "")).strip() if col_cont_resp else ""
                if not nome_resp:
                    log(f"Menor de idade sem responsável informado: '{nome_al}'", linha_planilha)
            elif maior is True:
                # Maior → campos de responsável ficam em branco
                pass
            else:
                log(f"Não foi possível determinar maioridade de '{nome_al}' (data ausente/inválida)", linha_planilha)

        # --- Dados bancários ---
        nome_banco = codigo_banco = agencia = dig_ag = conta = dig_cc = ""
        orig_banco = orig_agencia = orig_dig_ag = orig_conta = orig_dig_cc = ""
        status_banco  = 0
        excecao_banco = ""
        laranja_banco = {"ag": False, "da": False, "cc": False, "dc": False}
        if row_bk is not None:
            cj = _escolher_conjunto(row_bk, conjuntos_bancarios)
            # Valores originais (sem tratamento) para as colunas de auditoria
            def _orig(col):
                v = str(row_bk.get(col, "")).strip() if col else ""
                return "" if v.lower() in ("nan", "") else v
            orig_banco   = _orig(cj["banco"])
            orig_agencia = _orig(cj["agencia"])
            orig_dig_ag  = _orig(cj["dig_ag"])
            orig_conta   = _orig(cj["conta"])
            orig_dig_cc  = _orig(cj["dig_cc"])

            # Banco
            banco_raw = _orig(cj["banco"])
            codigo_banco, nome_banco, confianca, alerta_banco = resolver_banco(banco_raw)
            if alerta_banco:
                log(alerta_banco, linha_planilha)

            # Valores brutos para AG, DA, CC, DC
            ag_raw  = _orig(cj["agencia"])
            da_raw  = _orig(cj["dig_ag"])
            cc_raw  = _orig(cj["conta"])
            dc_raw  = _orig(cj["dig_cc"])

            # ── Tratamento completo AG / DA / CC / DC ──────────────────────
            agencia, dig_ag, conta, dig_cc, status_banco, excecao_banco, laranja_banco =                 tratar_dados_bancarios(ag_raw, da_raw, cc_raw, dc_raw, codigo_banco)

            if excecao_banco:
                log(excecao_banco, linha_planilha)

        else:
            status_banco  = 0
            excecao_banco = ""
            laranja_banco = {"ag": False, "da": False, "cc": False, "dc": False}

        # --- Montar linha de saída ---
        # Datas como datetime nativo para que o numfmt mm-dd-yy do template funcione
        linhas_saida.append({
            "No.":       len(linhas_saida) + 1,
            "ID":            id_aluno,
            "TURMA":         turma,
            "NOME COMPLETO": nome_al,
            "Data Inicio":   None,
            "Data Final":    None,
            "E-MAIL":        email,
            "CPF":           cpf,
            "RG":            rg,
            "DATA NASCIMENTO": data_nasc if data_nasc else None,
            "ENDERECO COMPLETO (Logradouro-Bairro-Cidade)": endereco_completo,
            "CEP":           cep,
            "Nome Responsavel": nome_resp,
            "CPF - Respon.": cpf_resp,
            "E-mail Respon.": email_resp,
            "Contato Respon.": cont_resp,
            "Nome do Banco": (f"{codigo_banco.zfill(3)} - {nome_banco}"
                             if codigo_banco else nome_banco),
            "Agencia":       agencia,
            "Dig.Ag":        dig_ag,
            "Conta":         conta,
            "Dig.C/C":       dig_cc,
            # Validação bancária
            "Status":        status_banco,
            "Excecoes":      excecao_banco,
            # Metadados de cor laranja por campo bancário
            "_laranja_ag":   laranja_banco.get("ag", False),
            "_laranja_da":   laranja_banco.get("da", False),
            "_laranja_cc":   laranja_banco.get("cc", False),
            "_laranja_dc":   laranja_banco.get("dc", False),
            # Colunas de auditoria — dados brutos originais do formulário
            "Orig.Banco":    orig_banco   if orig_banco   not in ("nan", "")  else "",
            "Orig.Agencia":  orig_agencia if orig_agencia not in ("nan", "")  else "",
            "Orig.Dig.Ag":   orig_dig_ag  if orig_dig_ag  not in ("nan", "")  else "",
            "Orig.Conta":    orig_conta   if orig_conta   not in ("nan", "")  else "",
            "Orig.Dig.C/C":  orig_dig_cc  if orig_dig_cc  not in ("nan", "")  else "",
        })

    df_saida = pd.DataFrame(linhas_saida)
    return df_saida, logs


# ==============================
# EXPORTAÇÃO PARA XLSX
# ==============================

# Mapeamento: chave do dicionário de saída → nome da coluna na Tabela1
# (preserva os nomes exatos do template, incluindo acentos)
MAPA_COLUNAS = {
    "No.":       "No.",
    "ID":            "ID",
    "TURMA":         "TURMA",
    "NOME COMPLETO": "NOME COMPLETO",
    "Data Inicio":   "Data Início",
    "Data Final":    "Data Final",
    "E-MAIL":        "E-MAIL",
    "CPF":           "CPF",
    "RG":            "RG",
    "DATA NASCIMENTO": "DATA NASCIMENTO",
    "ENDERECO COMPLETO (Logradouro-Bairro-Cidade)": "ENDEREÇO COMPLETO (Logradouro-Bairro-Cidade)",
    "CEP":           "CEP",
    "Nome Responsavel": "Nome Responsável",
    "CPF - Respon.": "CPF - Respon.",
    "E-mail Respon.": "E-mail Respon.",
    "Contato Respon.": "Contato Respon.",
    "Nome do Banco": "Nome do Banco",
    "Agencia":       "Agência",
    "Dig.Ag":        "Díg.Ag",
    "Conta":         "Conta",
    "Dig.C/C":       "Díg.C/C",
    # Colunas de validação e auditoria (fora da Tabela1, logo após a última coluna)
    "Status":        "Status",
    "Excecoes":      "Exceções do algoritmo",
    "Orig.Banco":    "Banco (original)",
    "Orig.Agencia":  "Agência (original)",
    "Orig.Dig.Ag":   "Díg.Ag (original)",
    "Orig.Conta":    "Conta (original)",
    "Orig.Dig.C/C":  "Díg.C/C (original)",
}


def exportar_xlsx(df_saida, path_saida, path_template=None):
    """
    Copia o template, limpa a linha-modelo de dados (linha 2), insere as
    linhas novas clonando os estilos dessa linha-modelo, e expande o ref
    da Tabela1 para cobrir todas as linhas inseridas.

    Se o template não existir, emite um xlsx simples sem formatação.
    """
    import shutil

    n = len(df_saida)

    # ── COM TEMPLATE ────────────────────────────────────────────────────────
    if path_template and os.path.exists(path_template):
        shutil.copy2(path_template, path_saida)
        wb = load_workbook(path_saida)
        ws = wb.active

        # Localiza a Tabela1 e descobre a linha do cabeçalho
        tabela = ws.tables.get("Tabela1")
        if tabela is None:
            raise Exception("Tabela1 não encontrada no template.")

        # ref da tabela: ex. "A1:U2" → linha_header=1, col_ini=A, col_fim=U
        from openpyxl.utils import range_boundaries, get_column_letter
        col_min, row_header, col_max, _ = range_boundaries(tabela.ref)
        linha_modelo = row_header + 1          # linha 2: dados de exemplo
        linha_dados_ini = linha_modelo          # vamos escrever a partir daqui

        # Lê os estilos da linha-modelo antes de qualquer modificação
        estilos_modelo = {}
        for c_idx in range(col_min, col_max + 1):
            origem = ws.cell(row=linha_modelo, column=c_idx)
            estilos_modelo[c_idx] = {
                "font":          copy(origem.font),
                "fill":          copy(origem.fill),
                "border":        copy(origem.border),
                "alignment":     copy(origem.alignment),
                "number_format": origem.number_format,
            }

        # Monta mapa: nome_coluna_template → índice de coluna (1-based)
        cabecalho_para_col = {}
        for c_idx in range(col_min, col_max + 1):
            nome = ws.cell(row=row_header, column=c_idx).value
            if nome:
                cabecalho_para_col[nome] = c_idx

        # Limpa o conteúdo da linha-modelo (valor e hyperlink), mantém estilos
        for c_idx in range(col_min, col_max + 1):
            cell = ws.cell(row=linha_modelo, column=c_idx)
            cell.value     = None
            cell.hyperlink = None

        # Insere linhas extras se n > 1 (a linha 2 já existe)
        if n > 1:
            ws.insert_rows(linha_modelo + 1, n - 1)

        # Chaves que vão dentro da Tabela1 e chaves de auditoria (fora)
        CHAVES_AUDITORIA = ["Status", "Excecoes",
                            "Orig.Banco", "Orig.Agencia", "Orig.Dig.Ag",
                            "Orig.Conta", "Orig.Dig.C/C"]
        MAPA_AUDITORIA   = {k: MAPA_COLUNAS[k] for k in CHAVES_AUDITORIA}
        MAPA_TABELA      = {k: v for k, v in MAPA_COLUNAS.items()
                            if k not in CHAVES_AUDITORIA}

        # Estilo para as colunas de auditoria (fundo cinza, borda fina, texto menor)
        fill_audit  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        font_audit  = Font(name="Calibri", size=9, italic=True, color="444444")
        border_audit = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        align_audit = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # Cabeçalhos das colunas de auditoria (linha 1, após col_max)
        col_audit_ini = col_max + 1
        for i, (chave_aud, nome_cab) in enumerate(MAPA_AUDITORIA.items()):
            c = col_audit_ini + i
            cell = ws.cell(row=row_header, column=c, value=nome_cab)
            cell.font      = Font(name="Calibri", size=9, bold=True, color="444444")
            cell.fill      = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
            cell.border    = border_audit
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Exceções: coluna mais larga para texto descritivo
            ws.column_dimensions[get_column_letter(c)].width = (
                6 if chave_aud == "Status" else
                45 if chave_aud == "Excecoes" else 18
            )

        # Escreve dados: Tabela1 + auditoria
        for r_offset, (_, row_df) in enumerate(df_saida.iterrows()):
            r = linha_dados_ini + r_offset
            # Mapa: chave do df → flag de laranja correspondente
            LARANJA_MAP = {
                "Agencia": "_laranja_ag",
                "Dig.Ag":  "_laranja_da",
                "Conta":   "_laranja_cc",
                "Dig.C/C": "_laranja_dc",
            }

            # Colunas da Tabela1
            for chave_df, nome_template in MAPA_TABELA.items():
                c_idx = cabecalho_para_col.get(nome_template)
                if c_idx is None:
                    continue
                val = row_df.get(chave_df, None)
                if isinstance(val, str) and val.strip() == "":
                    val = None
                cell = ws.cell(row=r, column=c_idx, value=val)
                est = estilos_modelo[c_idx]
                cell.font          = copy(est["font"])
                cell.border        = copy(est["border"])
                cell.alignment     = copy(est["alignment"])
                cell.number_format = est["number_format"]
                # Aplicar fill laranja se o campo bancário tem inconsistência
                flag_col = LARANJA_MAP.get(chave_df)
                if flag_col and row_df.get(flag_col, False):
                    cell.fill = copy(FILL_LARANJA)
                else:
                    cell.fill = copy(est["fill"])
            # Colunas de auditoria (fora da tabela)
            for i, chave_df in enumerate(MAPA_AUDITORIA.keys()):
                c = col_audit_ini + i
                val = row_df.get(chave_df, None)
                if isinstance(val, str) and val.strip() in ("", "nan"):
                    val = None
                # Status é numérico; Excecoes e originais são texto (@)
                if chave_df == "Status":
                    val = int(val) if val not in (None, "") else None
                cell = ws.cell(row=r, column=c, value=val)
                cell.font      = copy(font_audit)
                cell.fill      = copy(fill_audit)
                cell.border    = copy(border_audit)
                cell.alignment = copy(align_audit)
                cell.number_format = "0" if chave_df == "Status" else "@"

        # Atualiza o ref da Tabela1 (não inclui as colunas de auditoria)
        ultima_linha = linha_dados_ini + n - 1
        col_ini_letra = get_column_letter(col_min)
        col_fim_letra = get_column_letter(col_max)
        tabela.ref = f"{col_ini_letra}{row_header}:{col_fim_letra}{ultima_linha}"

        wb.save(path_saida)

    # ── SEM TEMPLATE (fallback) ─────────────────────────────────────────────
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados Bancários"
        ws.sheet_view.showGridLines = False

        cabecalhos = list(MAPA_COLUNAS.values())
        borda = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for c_idx, nome in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=c_idx, value=nome)
            cell.font      = Font(name="Calibri", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.border    = borda

        for r_offset, (_, row_df) in enumerate(df_saida.iterrows(), start=2):
            for c_idx, chave_df in enumerate(MAPA_COLUNAS.keys(), 1):
                val = row_df.get(chave_df, None)
                if isinstance(val, str) and val.strip() == "":
                    val = None
                cell = ws.cell(row=r_offset, column=c_idx, value=val)
                cell.font   = Font(name="Calibri", size=11)
                cell.border = borda

        wb.save(path_saida)


def salvar_logs(logs, path_saida_xlsx):
    """Salva error_logs.txt no mesmo diretório da planilha de saída."""
    if not logs:
        return None
    dir_saida = os.path.dirname(path_saida_xlsx)
    path_log = os.path.join(dir_saida, "error_logs.txt")
    with open(path_log, "w", encoding="utf-8") as f:
        f.write(f"Registro de exceções — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")
        for linha in logs:
            f.write(linha + "\n")
    return path_log


# ==============================
# WORKER THREAD (não trava a UI)
# ==============================

class WorkerThread(QThread):
    concluido = Signal(object, list)
    erro = Signal(str)

    def __init__(self, path_alunos, path_bancarios, path_template):
        super().__init__()
        self.path_alunos = path_alunos
        self.path_bancarios = path_bancarios
        self.path_template = path_template

    def run(self):
        try:
            df, logs = processar(self.path_alunos, self.path_bancarios, self.path_template)
            self.concluido.emit(df, logs)
        except Exception as e:
            self.erro.emit(str(e))


# ==============================
# INTERFACE GRÁFICA (PySide6)
# ==============================

def resource_path(rel):
    try:
        base = sys._MEIPASS
    except Exception:
        base = os.path.abspath(".")
    return os.path.join(base, rel)


BUTTON_STYLE = """
QPushButton {
    background-color: #f9b02e;
    color: black;
    border: none;
    border-radius: 8px;
    padding: 10px;
    font-weight: bold;
    font-size: 14px;
}
QPushButton:hover { background-color: #ffd166; }
QPushButton:pressed { background-color: #e69500; }
QPushButton:disabled { background-color: #666666; color: #aaaaaa; }
"""

REMOVE_STYLE = """
QPushButton {
    background-color: #444444;
    color: #cccccc;
    border: none;
    border-radius: 8px;
    padding: 10px;
    font-size: 12px;
}
QPushButton:hover { background-color: #666666; }
"""


class ToggleSwitch(QCheckBox):
    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setFixedHeight(28)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        track_rect = QRectF(0, 4, 44, 20)
        painter.setBrush(QColor("#f9b02e") if self.isChecked() else QColor("#3a3a3a"))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRoundedRect(track_rect, 10, 10)
        knob_x = 24 if self.isChecked() else 2
        painter.setBrush(QColor("white"))
        painter.drawEllipse(QRectF(knob_x, 6, 16, 16))
        painter.setPen(QColor("#dddddd"))
        painter.drawText(54, 19, self.text())


def criar_janela():
    app = QApplication(sys.argv)

    window = QWidget()
    window.setWindowTitle("BolsistaDB")
    window.setWindowIcon(QIcon(resource_path("bolsistadb.ico")))
    window.resize(960, 540)
    window.setMinimumSize(700, 460)
    window.setStyleSheet("""
        QWidget { background-color: #1e1e1e; font-family: "Segoe UI"; color: white; }
        QScrollArea { border: none; background: transparent; }
        QScrollBar:vertical { background: #2b2b2b; width: 10px; border-radius: 5px; }
        QScrollBar::handle:vertical { background: #f9b02e; border-radius: 5px; }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        QCheckBox { color: #dddddd; font-size: 12px; spacing: 12px; background: transparent; }
        QCheckBox::indicator { width: 42px; height: 22px; border-radius: 11px;
                               background-color: #3a3a3a; border: 1px solid #555555; }
        QCheckBox::indicator:checked { background-color: #f9b02e; border: 1px solid #f9b02e; }
    """)

    # Fundo bg_hbr.png (mesmo recurso do SmartPC)
    bg_label = QLabel(window)
    bg_pixmap = QPixmap(resource_path("bg_hbr.png"))
    bg_label.setPixmap(bg_pixmap)
    bg_label.setScaledContents(True)

    scroll = QScrollArea(window)
    scroll.setWidgetResizable(True)
    scroll.setStyleSheet("background: transparent;")

    container = QWidget()
    container.setStyleSheet("background: transparent;")
    scroll.setWidget(container)

    card = QFrame()
    card.setObjectName("card")
    card.setStyleSheet("""
        #card {
            background-color: rgba(30,30,30,220);
            border-radius: 18px;
            border: 1px solid rgba(255,255,255,25);
        }
    """)

    # Template resolvido automaticamente a partir da pasta do script/exe
    PATH_TEMPLATE = resource_path("template.xlsx")

    # Estado dos arquivos
    state = {
        "alunos": None,
        "bancarios": None,
        "worker": None,
        "df_resultado": None,
        "logs_resultado": None,
    }

    # --- Labels ---
    label_alunos    = QLabel("-")
    label_bancarios = QLabel("-")

    def estilo_label():
        return """
            background-color: #2b2b2b;
            border: 1px solid #3a3a3a;
            border-radius: 8px;
            padding: 8px;
            color: #cccccc;
            font-size: 11px;
        """

    for lbl in [label_alunos, label_bancarios]:
        lbl.setWordWrap(True)
        lbl.setStyleSheet(estilo_label())
        lbl.setMinimumHeight(24)
        lbl.setMaximumHeight(48)

    btn_executar = QPushButton("Gerar Planilha")
    btn_executar.setFixedSize(150, 44)
    btn_executar.setStyleSheet(BUTTON_STYLE)
    btn_executar.setEnabled(False)

    status_label = QLabel("")
    status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
    status_label.setStyleSheet("color: #aaaaaa; font-size: 11px; background: transparent;")

    def atualizar_status():
        ok = bool(state["alunos"]) and bool(state["bancarios"])
        btn_executar.setEnabled(ok)

    def criar_bloco(titulo, callback_sel, callback_rem, label_widget, aceita="Excel (*.xlsx *.xls)"):
        bloco = QFrame()
        bloco.setStyleSheet("""
            QFrame { background-color: rgba(255,255,255,8); border-radius: 12px; }
        """)
        layout = QVBoxLayout(bloco)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(8)

        titulo_lbl = QLabel(titulo)
        titulo_lbl.setStyleSheet("color: white; font-size: 13px; font-weight: bold; background: transparent;")
        layout.addWidget(titulo_lbl)

        btn_row = QHBoxLayout()
        btn_sel = QPushButton("Selecionar")
        btn_sel.setFixedSize(110, 32)
        btn_sel.setStyleSheet(BUTTON_STYLE)
        btn_sel.clicked.connect(callback_sel)

        btn_rem = QPushButton("Remover")
        btn_rem.setFixedSize(110, 32)
        btn_rem.setStyleSheet(REMOVE_STYLE)
        btn_rem.clicked.connect(callback_rem)

        btn_row.addWidget(btn_sel)
        btn_row.addWidget(btn_rem)
        btn_row.addStretch()
        layout.addLayout(btn_row)
        layout.addWidget(label_widget)
        return bloco

    def sel_alunos():
        f, _ = QFileDialog.getOpenFileName(window, "Selecionar lista de bolsistas", "", "Excel (*.xlsx *.xls)")
        if f:
            state["alunos"] = f
            label_alunos.setText(os.path.basename(f))
        atualizar_status()

    def rem_alunos():
        state["alunos"] = None
        label_alunos.setText("-")
        atualizar_status()

    def sel_bancarios():
        f, _ = QFileDialog.getOpenFileName(window, "Selecionar planilha de dados bancários", "", "Excel (*.xlsx *.xls)")
        if f:
            state["bancarios"] = f
            label_bancarios.setText(os.path.basename(f))
        atualizar_status()

    def rem_bancarios():
        state["bancarios"] = None
        label_bancarios.setText("-")
        atualizar_status()

    def ao_concluir(df, logs):
        state["df_resultado"]   = df
        state["logs_resultado"] = logs

        btn_executar.setEnabled(True)
        status_label.setText("")

        n_logs = len(logs)
        msg = f"Processamento concluído com {len(df)} alunos."
        if n_logs:
            msg += f"\n{n_logs} aviso(s) registrado(s) em error_logs.txt."

        path_saida, _ = QFileDialog.getSaveFileName(
            window, "Salvar planilha de saída", "saida_dados_bancarios.xlsx", "Excel (*.xlsx)"
        )
        if not path_saida:
            QMessageBox.warning(window, "Cancelado", "Arquivo não salvo.")
            return

        try:
            exportar_xlsx(df, path_saida, PATH_TEMPLATE)
            path_log = salvar_logs(logs, path_saida) if logs else None
            detalhe = f"Planilha salva em:\n{path_saida}"
            if path_log:
                detalhe += f"\n\nLog de exceções salvo em:\n{path_log}"
            QMessageBox.information(window, "Concluído", msg + "\n\n" + detalhe)
            if abrir_checkbox.isChecked():
                try:
                    os.startfile(path_saida)
                except Exception:
                    subprocess.call(["open", path_saida])
        except Exception as e:
            QMessageBox.critical(window, "Erro ao salvar", str(e))

    def ao_erro(msg):
        btn_executar.setEnabled(True)
        status_label.setText("")
        QMessageBox.critical(window, "Erro no processamento", msg)

    def executar():
        btn_executar.setEnabled(False)
        status_label.setText("Processando…")
        worker = WorkerThread(state["alunos"], state["bancarios"], PATH_TEMPLATE)
        worker.concluido.connect(ao_concluir)
        worker.erro.connect(ao_erro)
        state["worker"] = worker
        worker.start()

    btn_executar.clicked.connect(executar)

    # --- Layout do card ---
    card_layout = QVBoxLayout(card)
    card_layout.setContentsMargins(28, 28, 28, 28)
    card_layout.setSpacing(14)

    titulo = QLabel("BolsistaDB")
    titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
    titulo.setStyleSheet("""
        font-family: "Bahnschrift Condensed";
        font-size: 38px;
        font-weight: bold;
        color: white;
        padding-bottom: 2px;
        background: transparent;
    """)
    card_layout.addWidget(titulo)

    # subtitulo = QLabel("Automação de tratamento de dados bancários — HBR")
    # subtitulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
    # subtitulo.setStyleSheet("color: #888888; font-size: 11px; background: transparent; margin-bottom: 6px;")
    # card_layout.addWidget(subtitulo)

    card_layout.addWidget(criar_bloco("Lista de Bolsistas", sel_alunos, rem_alunos, label_alunos))
    card_layout.addWidget(criar_bloco("Planilha de Dados Bancários", sel_bancarios, rem_bancarios, label_bancarios))

    abrir_checkbox = ToggleSwitch("Abrir planilha quando estiver pronta")
    card_layout.addWidget(abrir_checkbox)

    exec_row = QHBoxLayout()
    exec_row.addStretch()
    exec_row.addWidget(btn_executar)
    exec_row.addStretch()
    card_layout.addLayout(exec_row)
    card_layout.addWidget(status_label)

    # --- Layout principal ---
    main_layout = QVBoxLayout(container)
    main_layout.addStretch()
    main_layout.addWidget(card, alignment=Qt.AlignmentFlag.AlignCenter)
    main_layout.addStretch()
    main_layout.setContentsMargins(30, 30, 30, 30)

    # --- Assinatura ---
    github_link = QLabel()
    github_link.setText(
        '<a href="https://github.com/imbaTIMvel/bolsistadb">'
        'BolsistaDB v0.1.0 - GitHub'
        '</a>'
        )
    github_link.setAlignment(Qt.AlignmentFlag.AlignCenter)
    github_link.setOpenExternalLinks(True)
    github_link.setStyleSheet("""
        QLabel {
            background-color: transparent;
            color: rgba(255,255,255,120);
            font-size: 11px;
        }
        QLabel:hover {
            color: #f9b02e;
        }
        """)
    footer = QLabel("Desenvolvido por: Diretoria Administrativa Financeira — DAF")
    footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
    footer.setStyleSheet("""
        QLabel {
            background-color: transparent;
            color: rgba(255,255,255,120);
            font-size: 10px;
            padding-bottom: 4px;
        }
    """)
    main_layout.addWidget(github_link)
    main_layout.addWidget(footer)

    window_layout = QVBoxLayout(window)
    window_layout.setContentsMargins(0, 0, 0, 0)
    window_layout.addWidget(scroll)

    def resize_event(event):
        bg_label.resize(window.size())
        scroll.resize(window.size())
        card.setMaximumWidth(620)

    window.resizeEvent = resize_event

    window.show()
    sys.exit(app.exec())


# ==============================
# ENTRY POINT
# ==============================

if __name__ == "__main__":
    criar_janela()
